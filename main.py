# подключаем библиотеки
import tkinter
from tkinter import *
from tkinter import messagebox, filedialog, ttk, Tk
from docxtpl import DocxTemplate
from openpyxl import load_workbook, Workbook

# подключаем графическую библиотеку
window = Tk()
# заголовок окна
window.title("Мой бухгалтер")

# Константы
SAVE_DIRECTORY = ''
EXCEL_FILE_FIRST = ''
EXCEL_FILE_SECOND = ''
EXCEL_FILE_THIRD = ''
UNIVERSE = 'СКЛЯРОВ ЛОХ'


def focus_next(event):
    """Выбирает следущий виджет интерфейса"""
    event.widget.tk_focusNext().focus()
    return "break"


def on_closing():
    """Обрабатывает закрытие окна"""
    if messagebox.askokcancel('Внимание', "Закрыть программу?"):
        window.destroy()


def clear_form(arr):
    """Очистить все поля ввода Entry"""
    global svo, target_direction
    svo = IntVar()
    target_direction = IntVar()
    for entry in arr:
        entry.delete(0, tkinter.END)


def error(info):
    """Выводит диалоговое окно с описание ошибк"""
    messagebox.askokcancel('Внимание', info)


def context():
    """Возвращает словарь со всеми переменными введенными в интерфейсе и их значения"""
    return {'reg_number': reg_number.get(),
            'surname': surname.get(),
            'name': name.get(),
            'patronymic': patronymic.get(),
            'date_birthday': date_birthday.get(),
            'snils': snils.get(),
            'inn': inn.get(),
            'citizenship': citizenship.get(),
            'id_doc': id_doc.get(),
            'series': series.get(),
            'number': number.get(),
            'date_id_doc': date_id_doc.get(),
            'office_doc': office_doc.get(),
            'address': address.get(),
            'tel_number': tel_number.get(),
            'spec_var_first': spec_var_first.get(),
            'spec_var_second': spec_var_second.get(),
            'spec_var_third': spec_var_third.get(),
            'parent_work': parent_work.get(),
            'certificate_score': certificate_score.get(),
            'form_education': form_education.get(),
            'svo': 'Да' if svo.get() else 'Нет',
            'target_direction': 'Да' if target_direction.get() else 'Нет',
            'parent_fio': parent_fio.get(),
            'parent_ser_num': parent_ser_num_pass.get(),
            'parent_pass_info': parent_pass_info.get(),
            'parent_address': parent_address.get(),
            'parent_number': parent_number.get(),
            'base_education': base_education.get(),
            'finance': finance.get()
            }


def fill_word():
    """Заполнить документ Word информацией введенной из интерфейса"""
    doc = None
    doc2 = None
    # Заполнение заявления
    # Проверка вида заполнения документа
    if choice.get() == profession:
        doc = DocxTemplate('patterns/prof.docx')
    elif choice.get() == specialty:
        doc = DocxTemplate('patterns/special.docx')
    elif choice.get() == specialty_with_exam:
        if (spec_var_first.get() == '54.02.01 Дизайн (по отраслям)'
                or spec_var_first.get() == '07.02.01 Архитектура'
                or spec_var_first.get() == '55.02.02 Анимация и анимационное кино (по видам)'):
            doc = DocxTemplate('patterns/sp_with_ex.docx')
        else:
            return error(
                'При выборе специальности с экзаменом необходимо выбрать одну из следующих специальностей:' +
                '\n07.02.01 Архитектура,\n54.02.01 Дизайн (по отраслям),' +
                '\n55.02.02 Анимация и анимационное кино (по видам)')
    if doc:
        doc.render(context())
        all_way = save_file('Создание ЗАЯВЛЕНИЯ word файла для заполнения')
        while not all_way:
            all_way = save_file('Создание ЗАЯВЛЕНИЯ word файла для заполнения')
        doc.save(all_way)

    # Заполнение  согласия на обработку персональных данных
    # Проверка вида заполнения документа
    if approval.get() == adult:
        doc2 = DocxTemplate('patterns/adult.docx')
    elif approval.get() == minor:
        doc2 = DocxTemplate('patterns/minor.docx')
    if doc2:
        doc2.render(context())

        all_way2 = save_file('Создание СОГЛАСИЯ word файла для заполнения')
        while not all_way2:
            all_way2 = save_file('Создание СОГЛАСИЯ word файла для заполнения')
        doc2.save(all_way2)
        error('Созданы заполненные word файлы!')


def save_file(message, type_file='word'):
    """Возвращает путь, имя файла и расширение, с которым его необходимо сохранить"""
    if type_file == 'excel':
        file_path = filedialog.asksaveasfilename(title=message, defaultextension='xlsx',
                                                 filetypes=[("Excel files", "*.xlsx")])
    else:
        file_path = filedialog.asksaveasfilename(title=message, defaultextension=".docx",
                                                 filetypes=[("Word files", "*.docx")])
    return file_path


def select_excel_file(ex_but='first'):
    '''Возвращает уже созданный файл excel для дальнейшего редактирования'''

    global EXCEL_FILE_FIRST
    global EXCEL_FILE_SECOND
    global EXCEL_FILE_THIRD
    global l22
    global l23
    global l33

    if ex_but == 'first':
        EXCEL_FILE_FIRST = filedialog.askopenfilename(title='Выбор РЕЙТИНГ excel файла для заполнения',
                                                      defaultextension='xlsx',
                                                      filetypes=[("Excel files", "*.xlsx")])
        if EXCEL_FILE_FIRST:
            # обновление строки состояния выбранного файла первого excel
            l22.config(text=EXCEL_FILE_FIRST)
            return EXCEL_FILE_FIRST

    elif ex_but == 'second':
        EXCEL_FILE_SECOND = filedialog.askopenfilename(title='Выбор ОБЩИЙ excel файла для заполнения',
                                                       defaultextension='xlsx',
                                                       filetypes=[("Excel files", "*.xlsx")])
        if EXCEL_FILE_SECOND:
            # обновление строки состояния выбранного файла второго excel
            l23.config(text=EXCEL_FILE_SECOND)
            return EXCEL_FILE_SECOND

    elif ex_but == 'third':
        EXCEL_FILE_THIRD = filedialog.askopenfilename(title='Выбор АИС excel файла для заполнения',
                                                      defaultextension='xlsx',
                                                      filetypes=[('Excel files', '*.xlsx')])
        if EXCEL_FILE_THIRD:
            # обновление строки состояния выбранного файла третьего excel
            l33.config(text=EXCEL_FILE_THIRD)
            return EXCEL_FILE_THIRD


def create_excel_file(ex_but='first'):
    '''Создает новый excel файл и возвращает его для дальнейшего редактирования'''

    global EXCEL_FILE_FIRST
    global EXCEL_FILE_SECOND
    global EXCEL_FILE_THIRD
    global l22
    global l23
    global l33

    work_book = Workbook()

    if ex_but == 'first':
        EXCEL_FILE_FIRST = save_file('Создание РЕЙТИНГ excel файла для заполнения', type_file='excel')
        if EXCEL_FILE_FIRST:
            # обновление строки состояния выбранного файла
            l22.config(text=EXCEL_FILE_FIRST)
            work_book.save(EXCEL_FILE_FIRST)
            return EXCEL_FILE_FIRST

    elif ex_but == 'second':
        EXCEL_FILE_SECOND = save_file('Создание ОБЩИЙ excel файла для заполнения', type_file='excel')
        if EXCEL_FILE_SECOND:
            # обновление строки состояния выбранного файла
            l23.config(text=EXCEL_FILE_SECOND)
            work_book.save(EXCEL_FILE_SECOND)
            return EXCEL_FILE_SECOND

    elif ex_but == 'third':
        EXCEL_FILE_THIRD = save_file('Создание АИС excel файла для заполнения', type_file='excel')
        if EXCEL_FILE_THIRD:
            # обновление строки состояния выбранного файла
            l33.config(text=EXCEL_FILE_THIRD)
            work_book.save(EXCEL_FILE_THIRD)
            return EXCEL_FILE_THIRD


def fill_excel():
    '''Заполнить документ excel информацией введенной из интерфейса'''
    global EXCEL_FILE_FIRST
    global EXCEL_FILE_SECOND
    global EXCEL_FILE_THIRD

    # Проверка выбраны ли существующие файлы для заполнения
    while not EXCEL_FILE_FIRST:
        EXCEL_FILE_FIRST = select_excel_file(ex_but='first')

    while not EXCEL_FILE_SECOND:
        EXCEL_FILE_SECOND = select_excel_file(ex_but='second')

    while not EXCEL_FILE_THIRD:
        EXCEL_FILE_THIRD = select_excel_file(ex_but='third')

    '''Заполняем первый excel'''
    # Открываем существующий файл первый Excel
    work_book_first = load_workbook(EXCEL_FILE_FIRST)
    e = work_book_first.active

    #  Заполняем шапку excel
    head_excel_first = ['Номер заявления',
                        'ФИО абитуриента',
                        'Оригинал/копия аттестата',
                        'Средний балл',
                        'Специальность (1)',
                        'Специальность (2)',
                        'Специальность (3)',
                        'Форма обучения']

    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']

    if e['A1'].value is None:
        for i in range(8):
            e[f'{alphabet[i]}1'] = head_excel_first[i]

    #  Находим первую пустую строку в столбце "B"
    empty_row = 1
    while e[f'B{empty_row}'].value is not None:
        empty_row += 1

    # Записываем данные в пустую строку
    data = context()
    e[f'B{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
    e[f'D{empty_row}'] = data['certificate_score']
    e[f'E{empty_row}'] = data['spec_var_first']
    e[f'F{empty_row}'] = data['spec_var_second']
    e[f'G{empty_row}'] = data['spec_var_third']
    e[f'H{empty_row}'] = data['form_education']

    work_book_first.save(EXCEL_FILE_FIRST)
    work_book_first.close()

    '''Заполняем второй excel'''
    # Открываем существующий файл второй Excel
    work_book_second = load_workbook(EXCEL_FILE_SECOND)
    e2 = work_book_second.active

    #  Заполняем шапку excel
    head_excel_second = ['Регистрационный номер',
                         'Фамилия',
                         'Имя',
                         'Отчество',
                         'Дата рождения',
                         'СНИЛС',
                         'ИНН',
                         'Гражданство',
                         'Документ, удостоверяющий личность',
                         'Серия',
                         'Номер',
                         'Кем выдан',
                         'Когда выдан',
                         'Проживающий по адресу',
                         'Телефон',
                         'Специальность 1',
                         'Сведения о родителях',
                         'Средний балл аттестата',
                         'Форма обучения',
                         'Участник СВО',
                         'Целевое направление']

    if e2['A1'].value is None:
        for i in range(21):
            e2[f'{alphabet[i]}1'] = head_excel_second[i]

    #  Находим первую пустую строку в столбце "B"
    empty_row_sec = 1
    while e2[f'B{empty_row_sec}'].value is not None:
        empty_row_sec += 1

    #  Записываем данные в пустую строку
    # Загружаем все данные из интерфейса как словарь
    data2 = context()
    # Удаляем вторую и третью специальность
    del data2['spec_var_second']
    del data2['spec_var_third']
    # Переводим словарь в список
    data2 = [val for val in data2.values()]
    # Объединяем все сведения о родителях
    data2[16] = ' '.join(data2[21:]) + ' ' + data2[16]
    # Заполняем пустую строку данными из списка
    for i in range(21):
        e2[f'{alphabet[i]}{empty_row_sec}'] = data2[i]

    work_book_second.save(EXCEL_FILE_SECOND)
    work_book_second.close()

    '''Заполняем третий excel'''
    # Открываем существующий файл третий Excel
    work_book_third = load_workbook(EXCEL_FILE_THIRD)
    e3 = work_book_third.active

    #  Заполняем шапку excel
    head_excel_third = ['Номер заявления',
                        '№ ЕПГУ',
                        'Фамилия абитуриента',
                        'Имя абитуриента',
                        'Имя абитуриента',
                        'Дата рождения',
                        'Серия удостоверяющего документа',
                        'Номер удостоверяющего документа',
                        'СНИЛС',
                        'Дата подачи заявления',
                        'Источник подачи заявления',
                        'Специальность (1)',
                        'Специальность (2)',
                        'Специальность (3)',
                        'Средний балл аттестата',
                        'Тип финансирования',
                        'Форма обучения',
                        'Базовое образование',
                        'Статус заявления',
                        'Статус специальности']

    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']

    if e3['A1'].value is None:
        for i in range(20):
            e3[f'{alphabet[i]}1'] = head_excel_third[i]

    #  Находим первую пустую строку в столбце "С"
    empty_row = 1
    while e3[f'C{empty_row}'].value is not None:
        empty_row += 1

    # Записываем данные в пустую строку
    data = context()
    e3[f'C{empty_row}'] = data['surname']
    e3[f'D{empty_row}'] = data['name']
    e3[f'E{empty_row}'] = data['patronymic']
    e3[f'F{empty_row}'] = data['date_birthday']
    e3[f'G{empty_row}'] = data['series']
    e3[f'H{empty_row}'] = data['number']
    e3[f'I{empty_row}'] = data['snils']
    e3[f'L{empty_row}'] = data['spec_var_first']
    e3[f'M{empty_row}'] = data['spec_var_second']
    e3[f'N{empty_row}'] = data['spec_var_third']
    e3[f'O{empty_row}'] = data['certificate_score']
    e3[f'P{empty_row}'] = data['finance']
    e3[f'Q{empty_row}'] = data['form_education']
    e3[f'R{empty_row}'] = data['base_education']

    work_book_third.save(EXCEL_FILE_THIRD)
    work_book_third.close()
    error('Все изменения в excel файлы внесены!')


# Cоздание надписи для полей ввода и размещение их по сетке

l1 = Label(window, text='Регистрационный номер')
l1.grid(row=0, column=0, sticky=E)

l2 = Label(window, text='Фамилия')
l2.grid(row=1, column=0, sticky=E)

l3 = Label(window, text='Имя')
l3.grid(row=2, column=0, sticky=E)

l4 = Label(window, text='Отчество')
l4.grid(row=3, column=0, sticky=E)

l5 = Label(window, text='Дата рождения')
l5.grid(row=4, column=0, sticky=E)

l6 = Label(window, text='СНИЛС')
l6.grid(row=5, column=0, sticky=E)

l7 = Label(window, text='ИНН')
l7.grid(row=6, column=0, sticky=E)

l8 = Label(window, text='Гражданство')
l8.grid(row=7, column=0, sticky=E)

l9 = Label(window, text='Документ, удостоверяющий личность')
l9.grid(row=8, column=0, sticky=E)

l10 = Label(window, text='Серия')
l10.grid(row=9, column=0, sticky=E)

l12 = Label(window, text='Номер')
l12.grid(row=11, column=0, sticky=E)

l13 = Label(window, text='Когда и кем выдан')
l13.grid(row=12, column=0, sticky=E)

l14 = Label(window, text='Проживающего(ей) по адресу')
l14.grid(row=13, column=0, sticky=E)

l15 = Label(window, text='Телефон')
l15.grid(row=14, column=0, sticky=E)

l16 = Label(window, text='Специальность 1')
l16.grid(row=15, column=0, sticky=E)

l17 = Label(window, text='Специальность 2')
l17.grid(row=16, column=0, sticky=E)

l18 = Label(window, text='Специальность 3')
l18.grid(row=17, column=0, sticky=E)

l19 = Label(window, text='Место работы родителя')
l19.grid(row=25, column=0, sticky=E)

l25 = Label(window, text='ФИО родителя')
l25.grid(row=19, column=0, sticky=E)

l26 = Label(window, text='Серия и номер паспорта родителя')
l26.grid(row=20, column=0, sticky=E)

l26 = Label(window, text='Когда и кем выдан паспорт родителя')
l26.grid(row=21, column=0, sticky=E)

l27 = Label(window, text='Адрес регистрации родителя')
l27.grid(row=22, column=0, sticky=E)

l28 = Label(window, text='Номер телефона родителя')
l28.grid(row=23, column=0, sticky=E)

l30 = Label(window, text='Тип финансирования:')
l30.grid(row=0, column=2)

l31 = Label(window, text='Базовое образование:')
l31.grid(row=6, column=2)

l21 = Label(window, text='Форма обучения:')
l21.grid(row=12, column=3)

l22 = Label(window, text='Шаблон документа заявления:')
l22.grid(row=0, column=3)

l29 = Label(window, text='Шаблон документа согласия:')
l29.grid(row=6, column=3)

l20 = Label(window, text='Средний балл аттестата')
l20.grid(row=26, column=0, sticky=E)

l21 = Label(window, text='Заполнение рейтинга')
l21.grid(row=29, column=0)

l22 = Label(window, text='')
l22.grid(row=29, column=1)

l24 = Label(window, text='Заполнение общего')
l24.grid(row=30, column=0)

l23 = Label(window, text='')
l23.grid(row=30, column=1)

l32 = Label(window, text='Заполнение аис')
l32.grid(row=31, column=0)

l33 = Label(window, text='')
l33.grid(row=31, column=1)

# Создание полей для ввода данных

reg_number = StringVar()
e1 = Entry(window, textvariable=reg_number, width=30)
e1.grid(row=0, column=1, sticky=W)

surname = StringVar()
e2 = Entry(window, textvariable=surname, width=30)
e2.grid(row=1, column=1, sticky=W)

name = StringVar()
e3 = Entry(window, textvariable=name, width=30)
e3.grid(row=2, column=1, sticky=W)

patronymic = StringVar()
e4 = Entry(window, textvariable=patronymic, width=30)
e4.grid(row=3, column=1, sticky=W)

date_birthday = StringVar()
e5 = Entry(window, textvariable=date_birthday, width=30)
e5.grid(row=4, column=1, sticky=W)

snils = StringVar()
e6 = Entry(window, textvariable=snils, width=30)
e6.grid(row=5, column=1, sticky=W)

inn = StringVar()
e7 = Entry(window, textvariable=inn, width=30)
e7.grid(row=6, column=1, sticky=W)

citizenship = StringVar()
e8 = Entry(window, textvariable=citizenship, width=30)
e8.grid(row=7, column=1, sticky=W)

id_doc = StringVar()
e9 = Entry(window, textvariable=id_doc, width=30)
e9.grid(row=8, column=1, sticky=W)

series = StringVar()
e10 = Entry(window, textvariable=series, width=30)
e10.grid(row=9, column=1, sticky=W)

number = StringVar()
e11 = Entry(window, textvariable=number, width=30)
e11.grid(row=11, column=1, sticky=W)

date_id_doc = StringVar()
e12 = Entry(window, textvariable=date_id_doc, width=25)
e12.grid(row=12, column=1, sticky=W)

office_doc = StringVar()
e13 = Entry(window, textvariable=office_doc, width=25)
e13.grid(row=12, column=1)

address = StringVar()
e14 = Entry(window, textvariable=address, width=60)
e14.grid(row=13, column=1, sticky=W)

tel_number = StringVar()
e15 = Entry(window, textvariable=tel_number, width=30)
e15.grid(row=14, column=1, sticky=W)

parent_fio = StringVar()
e16 = Entry(window, textvariable=parent_fio, width=45)
e16.grid(row=19, column=1, sticky=W)

parent_ser_num_pass = StringVar()
e17 = Entry(window, textvariable=parent_ser_num_pass, width=30)
e17.grid(row=20, column=1, sticky=W)

parent_pass_info = StringVar()
e18 = Entry(window, textvariable=parent_pass_info, width=60)
e18.grid(row=21, column=1, sticky=W)

parent_address = StringVar()
e19 = Entry(window, textvariable=parent_address, width=60)
e19.grid(row=22, column=1, sticky=W)

parent_number = StringVar()
e20 = Entry(window, textvariable=parent_number, width=30)
e20.grid(row=23, column=1, sticky=W)

parent_work = StringVar()
e21 = Entry(window, textvariable=parent_work, width=60)
e21.grid(row=25, column=1, sticky=W)

certificate_score = StringVar()
e22 = Entry(window, textvariable=certificate_score)
e22.grid(row=26, column=1, sticky=W)

# Выбор типа финансирования
budget = 'Бюджет'
commerce = 'Коммерция'

finance = StringVar(value=budget)

rb6 = Radiobutton(window, text=budget, value=budget, variable=finance)
rb6.grid(row=1, column=2)

rb7 = Radiobutton(window, text=commerce, value=commerce, variable=finance)
rb7.grid(row=2, column=2)

# Выбор базового образования
nine_cls = '9 классов'
elev_cls = '11 классов'

base_education = StringVar(value=nine_cls)

rb8 = Radiobutton(window, text=nine_cls, value=nine_cls, variable=base_education)
rb8.grid(row=7, column=2)

rb9 = Radiobutton(window, text=elev_cls, value=elev_cls, variable=base_education)
rb9.grid(row=8, column=2)

# Выбор шаблона заполнения word файла

profession = 'Профессия'
specialty = 'Специальность'
specialty_with_exam = 'Специальность c экзаменом'

choice = StringVar(value=profession)

rb1 = Radiobutton(window, text=profession, value=profession, variable=choice)
rb1.grid(row=1, column=3)

rb2 = Radiobutton(window, text=specialty, value=specialty, variable=choice)
rb2.grid(row=2, column=3)

rb3 = Radiobutton(window, text=specialty_with_exam, value=specialty_with_exam, variable=choice)
rb3.grid(row=3, column=3)

# Кнопки для выбора трех специальностей
specializations = ['38.02.08 Торговое дело',
                   '08.02.13 Монтаж и эксплуатация внутренних сантехнических устройств, кондиционирования воздуха и вентиляции',
                   '08.02.01 Строительство и эксплуатация зданий и сооружений',
                   '29.02.11 Полиграфическое производство',
                   '08.02.14 Эксплуатация и обслуживание многоквартирного дома',
                   '08.01.28 Мастер отделочных строительных и декоративных работ',
                   '54.01.20 Графический дизайнер',
                   '09.02.07 Информационные системы и программирование',
                   '42.02.02 Издательское дело',
                   '54.01.01 Исполнитель художественно - оформительских работ',
                   '54.02.01 Дизайн (по отраслям)',
                   '08.02.08 Монтаж и эксплуатация оборудования и систем газоснабжения',
                   '07.02.01 Архитектура',
                   '55.02.02 Анимация и анимационное кино (по видам)']

spec_var_first = StringVar(value=specializations[0])
spec_var_second = StringVar(value=specializations[0])
spec_var_third = StringVar(value=specializations[0])

combobox1 = ttk.Combobox(textvariable=spec_var_first, values=specializations, width=60)
combobox1.grid(row=15, column=1)

combobox1 = ttk.Combobox(textvariable=spec_var_second, values=specializations, width=60)
combobox1.grid(row=16, column=1)

combobox1 = ttk.Combobox(textvariable=spec_var_third, values=specializations, width=60)
combobox1.grid(row=17, column=1)

# Кнопки для выбора формы обучения

full_time = 'Очная'
corr_time = 'Заочная'

form_education = StringVar(value=full_time)

rb4 = Radiobutton(window, text=full_time, value=full_time, variable=form_education)
rb4.grid(row=13, column=3)

rb5 = Radiobutton(window, text=corr_time, value=corr_time, variable=form_education)
rb5.grid(row=14, column=3)

# Кнопки для выбора шаблона документа согласия

adult = 'Совершеннолетний'
minor = 'Несовершеннолетний'

approval = StringVar(value=adult)

rb4 = Radiobutton(window, text=adult, value=adult, variable=approval)
rb4.grid(row=7, column=3)

rb5 = Radiobutton(window, text=minor, value=minor, variable=approval)
rb5.grid(row=8, column=3)

# Кнопки для выбора целевого направления и участика сво

svo = IntVar()
target_direction = IntVar()

svo_check_but = Checkbutton(text='Участник СВО', variable=svo)
svo_check_but.grid(row=27, column=0)

target_direction_but = Checkbutton(text='Целевое направление', variable=target_direction)
target_direction_but.grid(row=27, column=1)

# Кнопки взаимодействия
ENTRIES = [eval(f'e{i}') for i in range(1, 23)]
clr_but = Button(window, text='Очистить форму', width=12, background='red', command=lambda: clear_form(ENTRIES))
clr_but.grid(row=20, column=3)

btn1 = Button(window, text="Заполнить Word", width=12, command=fill_word)
btn1.grid(row=22, column=3)

btn2 = Button(window, text="Заполнить Excel", width=12, command=fill_excel)
btn2.grid(row=23, column=3)

btn3 = Button(window, text='Выбрать файл', width=12, command=lambda: select_excel_file(ex_but='first'))
btn3.grid(row=29, column=2)

btn4 = Button(window, text='Создать новый', width=12, command=lambda: create_excel_file(ex_but='first'))
btn4.grid(row=29, column=3)

btn5 = Button(window, text='Выбрать файл', width=12, command=lambda: select_excel_file(ex_but='second'))
btn5.grid(row=30, column=2)

btn6 = Button(window, text='Создать новый', width=12, command=lambda: create_excel_file(ex_but='second'))
btn6.grid(row=30, column=3)

btn7 = Button(window, text='Выбрать файл', width=12, command=lambda: select_excel_file(ex_but='third'))
btn7.grid(row=31, column=2)

btn8 = Button(window, text='Создать новый', width=12, command=lambda: create_excel_file(ex_but='third'))
btn8.grid(row=31, column=3)

# Адаптиный дисплей
window.columnconfigure(tuple(range(3)), weight=1)
window.rowconfigure(tuple(range(31)), weight=1)

# Переход на новое поле ввода при нажатии Enter
window.bind_class('Entry', '<Return>', focus_next)

# сообщаем системе о том, что делать, когда окно закрывается
window.protocol("WM_DELETE_WINDOW", on_closing)

# пусть окно работает всё время до закрытия
window.mainloop()
