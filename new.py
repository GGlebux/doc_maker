import os
import sys

from PyQt6.QtCore import QDate
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QLabel,
    QLineEdit,
    QComboBox,
    QCheckBox,
    QPushButton,
    QDateEdit,
    QVBoxLayout,
    QHBoxLayout,
    QFileDialog,
    QMessageBox, QScrollArea, QFormLayout, QGroupBox, QRadioButton
)
from cx_Freeze import setup, Executable
from docxtpl import DocxTemplate
from openpyxl import load_workbook

# Константы
SAVE_DIRECTORY = ''
EXCEL_FILE_FIRST = ''
EXCEL_FILE_SECOND = ''
EXCEL_FILE_THIRD = ''
UNIVERSE = 'СКЛЯРОВ ЛОХ'

# Глобальные переменные
specializations = ['38.02.08 Торговое дело',
                   '08.02.13 Монтаж и эксплуатация внутренних сантехнических устройств, кондиционирования воздуха и вентиляции',
                   '08.02.01 Строительство и эксплуатация зданий и сооружений',
                   '29.02.11 Полиграфическое производство',
                   '08.02.14 Эксплуатация и обслуживание многоквартирного дома',
                   '08.01.28 Мастер отделочных строительных и декоративных работ',
                   '54.01.20 Графический дизайнер',
                   '09.02.07 Информационные системы и программирование',
                   '42.02.02 Издательское дело',
                   '54.01.01 Исполнитель художественно - оформительских работ',
                   '54.02.01 Дизайн (по отраслям)',
                   '08.02.08 Монтаж и эксплуатация оборудования и систем газоснабжения',
                   '07.02.01 Архитектура',
                   '55.02.02 Анимация и анимационное кино (по видам)']


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("приЁмка")
        s


    def clear_form(self):
        """Очистить все поля ввода Entry"""
        self.reg_number.clear()
        self.surname.clear()
        self.name.clear()
        self.patronymic.clear()
        self.date_birthday.setDate(QDate.currentDate())
        self.snils.clear()
        self.inn.clear()
        self.citizenship.clear()
        self.id_doc.clear()
        self.series.clear()
        self.number.clear()
        self.date_id_doc.setDate(QDate.currentDate())
        self.office_doc.clear()
        self.address.clear()
        self.tel_number.clear()
        self.spec_var_first.setCurrentIndex(0)
        self.spec_var_third.setCurrentIndex(0)
        self.base_education.setCurrentIndex(0)
        self.parent_work.clear()
        self.certificate_score.clear()
        self.form_education.setCurrentIndex(0)
        self.parent_fio.clear()
        self.parent_ser_num_pass.clear()
        self.parent_pass_info.clear()
        self.parent_address.clear()
        self.parent_number.clear()
        self.svo_checkbox.setChecked(False)
        self.target_direction_checkbox.setChecked(False)

    def get_input_data(self):
        """Возвращает словарь со всеми переменными введенными в интерфейсе и их значения"""
        return {
            'reg_number': self.reg_number.text(),
            'surname': self.surname.text(),
            'name': self.name.text(),
            'patronymic': self.patronymic.text(),
            'date_birthday': self.date_birthday.date().toString("dd.MM.yyyy"),
            'snils': self.snils.text(),
            'inn': self.inn.text(),
            'citizenship': self.citizenship.text(),
            'id_doc': self.id_doc.text(),
            'series': self.series.text(),
            'number': self.number.text(),
            'date_id_doc': self.date_id_doc.date().toString("dd.MM.yyyy"),
            'office_doc': self.office_doc.text(),
            'address': self.address.text(),
            'tel_number': self.tel_number.text(),
            'spec_var_first': self.spec_var_first.currentText(),
            'spec_var_second': self.spec_var_third.currentText(),
            'spec_var_third': self.base_education.currentText(),
            'parent_work': self.parent_work.text(),
            'certificate_score': self.certificate_score.text(),
            'form_education': self.form_education.currentText(),
            'svo': "Да" if self.svo_checkbox.isChecked() else "Нет",
            'target_direction': "Да" if self.target_direction_checkbox.isChecked() else "Нет",
            'parent_fio': self.parent_fio.text(),
            'parent_ser_num': self.parent_ser_num_pass.text(),
            'parent_pass_info': self.parent_pass_info.text(),
            'parent_address': self.parent_address.text(),
            'parent_number': self.parent_number.text(),
            'base_education': self.base_education.currentText(),
            'finance': ""}

    def fill_word(self):
        """Заполнить документ Word информацией введенной из интерфейса"""
        doc = None
        doc2 = None

        # Заполнение заявления
        # Проверка вида заполнения документа
        if self.spec_var_first.currentText() == "54.02.01 Дизайн (по отраслям)":
            doc = doc_prof
        elif self.spec_var_first.currentText() == "07.02.01 Архитектура" or self.spec_var_first.currentText() == "55.02.02 Анимация и анимационное кино (по видам)":
            doc = doc_special
        elif self.spec_var_first.currentText() == "54.02.01 Дизайн (по отраслям)" or self.spec_var_first.currentText() == "07.02.01 Архитектура" or self.spec_var_first.currentText() == "55.02.02 Анимация и анимационное кино (по видам)":
            doc = doc_sp_with_ex
        else:
            QMessageBox.warning(self, "Ошибка",
                                "При выборе специальности с экзаменом необходимо выбрать одну из следующих специальностей:" +
                                "\n07.02.01 Архитектура,\n54.02.01 Дизайн (по отраслям)," +
                                "\n55.02.02 Анимация и анимационное кино (по видам)")
            return

        data = self.get_input_data()
        if doc:
            doc.render(data)
            filename, _ = QFileDialog.getSaveFileName(self, "Сохранить заявление", "Word Files (*.docx)")
            if filename:
                doc.save(filename)

        # Заполнение согласия на обработку персональных данных
        if self.adultRadio.isChecked():
            doc2 = doc_adult
        elif self.minorRadio.isChecked():
            doc2 = doc_minor
        if doc2:
            doc2.render(data)
            filename, _ = QFileDialog.getSaveFileName(self, "Сохранить согласие на обработку данных",
                                                      "Word Files (*.docx)")
            if filename:
                doc2.save(filename)
                QMessageBox.information(self, "Успешно", "Файлы Word успешно созданы!")

    def fill_excel(self):
        """Заполнить документ Excel информацией введенной из интерфейса"""
        global EXCEL_FILE_FIRST, EXCEL_FILE_SECOND, EXCEL_FILE_THIRD

        # Проверка выбраны ли существующие файлы для заполнения
        while not EXCEL_FILE_FIRST:
            EXCEL_FILE_FIRST = self.select_excel_file("Выберите первый файл Excel")

        while not EXCEL_FILE_SECOND:
            EXCEL_FILE_SECOND = self.select_excel_file("Выберите второй файл Excel")

        while not EXCEL_FILE_THIRD:
            EXCEL_FILE_THIRD = self.select_excel_file("Выберите третий файл Excel")

        '''Заполняем первый excel'''
        # Открываем существующий файл первый Excel
        work_book_first = load_workbook(EXCEL_FILE_FIRST)
        e = work_book_first.active

        #  Заполняем шапку excel
        head_excel_first = ['Номер заявления',
                            'ФИО абитуриента',
                            'Оригинал/копия аттестата',
                            'Средний балл',
                            'Специальность (1)',
                            'Специальность (2)',
                            'Специальность (3)',
                            'Форма обучения']

        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U',
                    'V', 'W', 'X', 'Y', 'Z']

        if e['A1'].value is None:
            for i in range(8):
                e[f'{alphabet[i]}1'] = head_excel_first[i]

        #  Находим первую пустую строку в столбце "B"
        empty_row = 1
        while e[f'B{empty_row}'].value is not None:
            empty_row += 1

        # Записываем данные в пустую строку
        data = self.get_input_data()
        e[f'B{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
        e[f'D{empty_row}'] = data['certificate_score']
        e[f'E{empty_row}'] = data['spec_var_first']
        e[f'F{empty_row}'] = data['spec_var_second']
        e[f'G{empty_row}'] = data['spec_var_third']
        e[f'H{empty_row}'] = data['form_education']

        work_book_first.save(EXCEL_FILE_FIRST)
        work_book_first.close()

        '''Заполняем второй excel'''
        # Открываем существующий файл второй Excel
        work_book_second = load_workbook(EXCEL_FILE_SECOND)
        e2 = work_book_second.active

        #  Заполняем шапку excel
        head_excel_second = ['Регистрационный номер',
                             'Фамилия',
                             'Имя',
                             'Отчество',
                             'Дата рождения',
                             'СНИЛС',
                             'ИНН',
                             'Гражданство',
                             'Документ, удостоверяющий личность',
                             'Серия',
                             'Номер',
                             'Кем выдан',
                             'Когда выдан',
                             'Проживающий по адресу',
                             'Телефон',
                             'Специальность 1',
                             'Сведения о родителях',
                             'Средний балл аттестата',
                             'Форма обучения',
                             'Участник СВО',
                             'Целевое направление']

        if e2['A1'].value is None:
            for i in range(21):
                e2[f'{alphabet[i]}1'] = head_excel_second[i]

            # Находим первую пустую строку в столбце "B"
            empty_row_sec = 1
            while e2[f'B{empty_row_sec}'].value is not None:
                empty_row_sec += 1

            # Проверяем, не существует ли уже такой же ФИО в столбце B
            fio = self.get_input_data()['surname'] + ' ' + self.get_input_data()['name'] + ' ' + self.get_input_data()[
                'patronymic']
            for row in range(2, empty_row_sec):
                if e2[f'B{row}'].value == fio:
                    QMessageBox.warning(self, "Ошибка", "Абитуриент с такими ФИО уже есть в таблице!")
                    return

            # Записываем данные в пустую строку (если ФИО уникально)
            # Загружаем все данные из интерфейса как словарь
            data2 = self.get_input_data()
            # Удаляем вторую и третью специальность
            del data2['spec_var_second']
            del data2['spec_var_third']
            # Переводим словарь в список
            data2 = [val for val in data2.values()]
            # Объединяем все сведения о родителях
            data2[16] = ' '.join(data2[21:]) + ' ' + data2[16]
            # Заполняем пустую строку данными из списка
            for i in range(21):
                e2[f'{alphabet[i]}{empty_row_sec}'] = data2[i]

            work_book_second.save(EXCEL_FILE_SECOND)
            work_book_second.close()

        '''Заполняем третий excel'''
        # Открываем существующий файл третий Excel
        work_book_third = load_workbook(EXCEL_FILE_THIRD)
        e3 = work_book_third.active

        #  Заполняем шапку excel
        head_excel_third = ['Номер заявления',
                            '№ ЕПГУ',
                            'Фамилия абитуриента',
                            'Имя абитуриента',
                            'Имя абитуриента',
                            'Дата рождения',
                            'Серия удостоверяющего документа',
                            'Номер удостоверяющего документа',
                            'СНИЛС',
                            'Дата подачи заявления',
                            'Источник подачи заявления',
                            'Специальность (1)',
                            'Специальность (2)',
                            'Специальность (3)',
                            'Средний балл аттестата',
                            'Тип финансирования',
                            'Форма обучения',
                            'Базовое образование',
                            'Статус заявления',
                            'Статус специальности']

        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                    'T', 'U',
                    'V', 'W', 'X', 'Y', 'Z']

        if e3['A1'].value is None:
            for i in range(20):
                e3[f'{alphabet[i]}1'] = head_excel_third[i]

        #  Находим первую пустую строку в столбце "С"
        empty_row = 1
        while e3[f'C{empty_row}'].value is not None:
            empty_row += 1

        work_book_third.save(EXCEL_FILE_THIRD)
        work_book_third.close()

    def select_excel_file(self, title):
        """Открывает диалоговое окно для выбора файла Excel"""
        global SAVE_DIRECTORY
        filename, _ = QFileDialog.getOpenFileName(self, title, SAVE_DIRECTORY, "Excel Files (*.xlsx)")
        if filename:
            SAVE_DIRECTORY = filename
            return filename
        return None

    def save_file(self, title):
        """Открывает диалоговое окно для сохранения файла"""
        global SAVE_DIRECTORY
        filename, _ = QFileDialog.getSaveFileName(self, title, SAVE_DIRECTORY, "Word Files (*.docx)")
        if filename:
            SAVE_DIRECTORY = filename
            return True
        return False


def load_template(template_name):
    template_path = os.path.join("patterns", template_name)
    doc = DocxTemplate(template_path)
    return doc


# setup(
#     name="Admission Office",
#     version="2.0",
#     description="My Awesome Application",
#     executables=[Executable("new.py", base="Win32GUI")],
#     # Включаем папку шаблонов
#     include_files=[("patterns", "patterns")], )

if __name__ == "__main__":
    doc_prof = load_template('prof.docx')
    doc_special = load_template('special.docx')
    doc_sp_with_ex = load_template('sp_with_ex.docx')
    doc_adult = load_template('adult.docx')
    doc_minor = load_template('minor.docx')

    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
