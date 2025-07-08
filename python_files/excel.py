import traceback

from PyQt6.QtCore import QDir
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl.reader.excel import load_workbook


def select_file_loop(obj, excel_name, button, flag=True):
    while not getattr(obj, excel_name) and flag:
        button.click()


class Excel:
    def __init__(self, parent, data):
        self.parent = parent
        self.data = data
        self.alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I',
                         'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                         'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

        self.rating_excel = None
        self.total_excel = None
        self.aic_excel = None
        self.stream_excel = None
        self.svo_excel = None
        self.dormitory_excel = None
        self.orphan_excel = None

        self.is_rating_done = False
        self.is_total_done = False
        self.is_aic_done = False
        self.is_stream_done = False
        self.is_svo_done = False
        self.is_dormitory_done = False
        self.is_orphan_done = False

        self.stream_flag = False
        self.svo_flag = False
        self.dormitory_flag = False
        self.orphan_flag = False

    def start_up(self):
        """Запуск заполнения"""
        try:
            # Валидация формы
            if not self.parent.validator.validate():
                return

            # Проверка выбраны ли существующие файлы для заполнения
            select_file_loop(self, 'rating_excel', self.parent.rating_button)
            select_file_loop(self, 'total_excel', self.parent.total_button)
            select_file_loop(self, 'aic_excel', self.parent.aic_button)
            select_file_loop(self, 'stream_excel', self.parent.stream_button, self.stream_flag)
            select_file_loop(self, 'svo_excel', self.parent.svo_button, self.svo_flag)
            select_file_loop(self, 'dormitory_excel', self.parent.dormitory_button, self.dormitory_flag)
            select_file_loop(self, 'orphan_excel', self.parent.orphan_button, self.orphan_flag)

            data = self.data.get_input_data()
            # Заполнение
            self.fill_rating_excel(data)
            self.fill_total_excel(data)
            self.fill_aic_excel(data)

            if self.stream_flag:
                self.fill_stream_excel(data)
            if self.svo_flag:
                self.fill_svo(data)
            if self.dormitory_flag:
                self.fill_dormitory(data)
            if self.orphan_flag:
                self.fill_orphan(data)

            if self.everything_is_done():
                self.success()

            self.set_everything_not_done()
        except PermissionError as e:
            self.parent.logger.warning(f'Ошибка при заполнении Excel: {e}')
            QMessageBox.warning(self.parent, "Ошибка",
                                "Закройте все окна Excel и повторите попытку\n(иначе данные не сохранятся)")
        except Exception as e:
            error_info = traceback.format_exc()
            self.parent.logger.error(f'Возникла непредвиденная ошибка при заполнении Excel: {e}\n{error_info}')
            QMessageBox.warning(self.parent, 'Критическая ошибка',
                                f'Возникла непредвиденная ошибка при заполнении Excel:\n(обратитесь к разработчику)')

    def fill_rating_excel(self, data):
        """Заполняет Рейтинг excel"""
        # Открываем Рейтинг существующий файл Excel
        wb = load_workbook(self.rating_excel)
        e = wb.active

        #  Заполняем шапку excel
        head_excel = ['Номер заявления',
                      'ФИО абитуриента',
                      'Оригинал/копия аттестата',
                      'Средний балл',
                      'Специальность (1)',
                      'Специальность (2)',
                      'Специальность (3)',
                      'Форма обучения',
                      'Оригинал аттестата']

        if e['A1'].value is None:
            for i in range(9):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        #  Находим первую пустую строку в столбце "B"
        empty_row = 1
        while e[f'B{empty_row}'].value is not None:
            empty_row += 1

        # Проверка на уникальность
        try:
            old_fio = e[f'B{empty_row - 1}'].value
            new_fio = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
            if not self.check_unique(old_fio, new_fio, '<Рейтинг>'):
                return
        except TypeError:
            self.no_data_warning(self.rating_excel)

        # Записываем данные в пустую строку
        e[f'A{empty_row}'] = data['reg_number']
        e[f'B{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
        e[f'D{empty_row}'] = data['certificate_score']
        e[f'E{empty_row}'] = data['spec_var_first']
        e[f'F{empty_row}'] = data['spec_var_second']
        e[f'G{empty_row}'] = data['spec_var_third']
        e[f'H{empty_row}'] = data['form_education']
        e[f'I{empty_row}'] = data['certificate']

        wb.save(self.rating_excel)
        wb.close()
        self.is_rating_done = True

    def fill_total_excel(self, data):
        """Заполняет Общий excel"""
        # Открываем ОБЩИЙ существующий файл Excel
        wb = load_workbook(self.total_excel)
        e = wb.active

        #  Заполняем шапку excel
        head_excel = ['Регистрационный номер',
                      'Фамилия',
                      'Имя',
                      'Отчество',
                      'Дата рождения',
                      'СНИЛС',
                      'ИНН',
                      'Гражданство',
                      'Документ, удостоверяющий личность',
                      'Серия',
                      'Номер',
                      'Кем выдан',
                      'Когда выдан',
                      'Проживающий по адресу',
                      'Телефон',
                      'Специальность 1',
                      'Сведения о родителях',
                      'Средний балл аттестата',
                      'Форма обучения',
                      'Участник СВО',
                      'Целевое направление',
                      'Бюджет',
                      'Коммерция']

        if e['A1'].value is None:
            for i in range(23):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        # Находим первую пустую строку в столбце "B"
        empty_row = 1
        while e[f'B{empty_row}'].value is not None:
            empty_row += 1

        # Проверка значений на уникальность
        try:
            old_pasport = e[f'J{empty_row - 1}'].value + e[f'K{empty_row - 1}'].value
            new_passport = data['series'] + data['number']
            if not self.check_unique(old_pasport, new_passport, '<Общий>'):
                return
        except TypeError:
            self.no_data_warning(self.total_excel)

        e[f'A{empty_row}'] = data['reg_number']
        e[f'B{empty_row}'] = data['surname']
        e[f'C{empty_row}'] = data['name']
        e[f'D{empty_row}'] = data['patronymic']
        e[f'E{empty_row}'] = data['date_birthday']
        e[f'F{empty_row}'] = data['snils']
        e[f'G{empty_row}'] = data['inn']
        e[f'H{empty_row}'] = data['citizenship']
        e[f'I{empty_row}'] = data['id_doc']
        e[f'J{empty_row}'] = data['series']
        e[f'K{empty_row}'] = data['number']
        e[f'L{empty_row}'] = data['office_doc']
        e[f'M{empty_row}'] = data['date_id_doc']
        e[f'N{empty_row}'] = data['address']
        e[f'O{empty_row}'] = data['tel_number']
        e[f'P{empty_row}'] = data['spec_var_first']
        e[f'Q{empty_row}'] = '; '.join([data['parent_fio'],
                                        data['parent_ser_num'],
                                        data['parent_pass_info'],
                                        data['parent_address'],
                                        data['parent_work'],
                                        data['parent_number']])
        e[f'R{empty_row}'] = data['certificate_score']
        e[f'S{empty_row}'] = data['form_education']
        e[f'T{empty_row}'] = data['svo']
        e[f'U{empty_row}'] = data['target_direction']
        e[f'V{empty_row}'] = data['finance']['budget']
        e[f'W{empty_row}'] = data['finance']['commerce']

        wb.save(self.total_excel)
        wb.close()
        self.is_total_done = True

    def fill_aic_excel(self, data):
        """Заполняет АИС excel"""
        # Открываем АИС существующий файл Excel
        wb = load_workbook(self.aic_excel)
        e = wb.active

        #  Заполняем шапку excel
        head_excel = ['Номер заявления',
                      '№ ЕПГУ',
                      'Фамилия абитуриента',
                      'Имя абитуриента',
                      'Отчество абитуриента',
                      'Дата рождения',
                      'Серия удостоверяющего документа',
                      'Номер удостоверяющего документа',
                      'СНИЛС',
                      'Дата подачи заявления',
                      'Источник подачи заявления',
                      'Специальность (1)',
                      'Специальность (2)',
                      'Специальность (3)',
                      'Средний балл аттестата',
                      'Тип финансирования',
                      'Форма обучения',
                      'Базовое образование',
                      'Статус заявления',
                      'Статус специальности']

        if e['A1'].value is None:
            for i in range(20):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        #  Находим первую пустую строку в столбце "С"
        empty_row = 1
        while e[f'C{empty_row}'].value is not None:
            empty_row += 1

        # Проверка значений на уникальность
        try:
            old_pasport = e[f'G{empty_row - 1}'].value + e[f'H{empty_row - 1}'].value
            new_passport = data['series'] + data['number']
            if not self.check_unique(old_pasport, new_passport, '<АИС>'):
                return
        except TypeError:
            self.no_data_warning(self.aic_excel)

        # Записываем данные в пустую строку
        e[f'A{empty_row}'] = data['reg_number']
        e[f'C{empty_row}'] = data['surname']
        e[f'D{empty_row}'] = data['name']
        e[f'E{empty_row}'] = data['patronymic']
        e[f'F{empty_row}'] = data['date_birthday']
        e[f'G{empty_row}'] = data['series']
        e[f'H{empty_row}'] = data['number']
        e[f'I{empty_row}'] = data['snils']
        e[f'L{empty_row}'] = data['spec_var_first']
        e[f'M{empty_row}'] = data['spec_var_second']
        e[f'N{empty_row}'] = data['spec_var_third']
        e[f'O{empty_row}'] = data['certificate_score']
        e[f'P{empty_row}'] = 'Бюджет' if data['finance']['budget'] == '+' else 'Коммерция'
        e[f'Q{empty_row}'] = data['form_education']
        e[f'R{empty_row}'] = data['base_education']

        wb.save(self.aic_excel)
        wb.close()
        self.is_aic_done = True

    def fill_stream_excel(self, data):
        """Заполняет Поток excel"""
        # Открываем ПОТОК существующий файл Excel
        wb = load_workbook(self.stream_excel)
        e = wb.active

        #  Заполняем шапку excel
        head_excel = ['ФИО абитуриента',
                      'Специальность 1',
                      'Поток']

        if e['A1'].value is None:
            for i in range(3):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        #  Находим первую пустую строку в столбце "A"
        empty_row = 1
        while e[f'A{empty_row}'].value is not None:
            empty_row += 1

        # Проверка значений на уникальность
        try:
            old_fio = e[f'A{empty_row - 1}'].value
            new_fio = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
            if not self.check_unique(old_fio, new_fio, '<Поток>'):
                return
        except TypeError:
            self.no_data_warning(self.stream_excel)

        # Записываем данные в пустую строку
        e[f'A{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
        e[f'B{empty_row}'] = data['spec_var_first']
        e[f'C{empty_row}'] = data['stream']

        wb.save(self.stream_excel)
        wb.close()
        self.is_stream_done = True

    def fill_svo(self, data):
        """Заполняет СВО excel"""
        # Открываем существующий СВО excel
        wb = load_workbook(self.svo_excel)
        e = wb.active

        # Заполняем шапку Excel
        head_excel = ['ФИО абитуриента',
                      'Специальность 1',
                      'СВО']

        if e['A1'].value is None:
            for i in range(3):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        empty_row = 1
        while e[f'A{empty_row}'].value is not None:
            empty_row += 1

        # Проверка значений на уникальность
        try:
            old_fio = e[f'A{empty_row - 1}'].value
            new_fio = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
            if not self.check_unique(old_fio, new_fio, '<СВО>'):
                return
        except TypeError:
            self.no_data_warning(self.svo_excel)

        e[f'A{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
        e[f'B{empty_row}'] = data['spec_var_first']
        e[f'C{empty_row}'] = data['svo']

        wb.save(self.svo_excel)
        wb.close()
        self.is_svo_done = True

    def fill_dormitory(self, data):
        """Заполняет ОБЩЕЖИТИЕ excel"""
        # Открываем существующий ОБЩЕЖИТИЕ excel
        wb = load_workbook(self.dormitory_excel)
        e = wb.active

        # Заполняем шапку Excel
        head_excel = ['ФИО абитуриента',
                      'Специальность 1',
                      'Пол']

        if e['A1'].value is None:
            for i in range(3):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        empty_row = 1
        while e[f'A{empty_row}'].value is not None:
            empty_row += 1

        # Проверка значений на уникальность
        try:
            old_fio = e[f'A{empty_row - 1}'].value
            new_fio = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
            if not self.check_unique(old_fio, new_fio, '<ОБЩЕЖИТИЕ>'):
                return
        except TypeError:
            self.no_data_warning(self.dormitory_excel)

        e[f'A{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
        e[f'B{empty_row}'] = data['spec_var_first']
        e[f'C{empty_row}'] = data['gender']

        wb.save(self.dormitory_excel)
        wb.close()
        self.is_dormitory_done = True

    def fill_orphan(self, data):
        """Заполняет СИРОТЫ excel"""
        # Открываем существующий СИРОТЫ excel
        wb = load_workbook(self.orphan_excel)
        e = wb.active

        # Заполняем шапку Excel
        head_excel = ['ФИО абитуриента',
                      'Специальность 1',
                      'Сирота']

        if e['A1'].value is None:
            for i in range(3):
                e[f'{self.alphabet[i]}1'] = head_excel[i]

        empty_row = 1
        while e[f'A{empty_row}'].value is not None:
            empty_row += 1

        # Проверка значений на уникальность
        try:
            old_fio = e[f'A{empty_row - 1}'].value
            new_fio = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
            if not self.check_unique(old_fio, new_fio, '<СИРОТА>'):
                return
        except TypeError:
            self.no_data_warning(self.orphan_excel)

        e[f'A{empty_row}'] = data['surname'] + ' ' + data['name'] + ' ' + data['patronymic']
        e[f'B{empty_row}'] = data['spec_var_first']
        e[f'C{empty_row}'] = data['orphan']

        wb.save(self.orphan_excel)
        wb.close()
        self.is_orphan_done = True

    def select_excel_file(self, title):
        """Открывает диалоговое окно для выбора файла Excel"""
        filename, _ = QFileDialog.getOpenFileName(self.parent, title, QDir().homePath(), "Excel Files (*.xlsx)")
        if filename:
            return filename
        return None

    def check_unique(self, old_value, new_value, table):
        """Проверяет есть ли в таблице данные, подобные новым"""
        if old_value == new_value:
            self.parent.logger.warning(f'Дубликат абитуриента: {old_value}')
            QMessageBox.warning(self.parent, "Ошибка", f"Такой абитуриент уже есть в таблице: {table}!")
            return False
        return True

    def everything_is_done(self):
        return (self.is_rating_done
                and self.is_total_done
                and self.is_aic_done
                and (self.is_stream_done == self.stream_flag)
                and (self.is_svo_done == self.svo_flag)
                and (self.is_dormitory_done == self.dormitory_flag)
                and (self.is_orphan_done == self.orphan_flag))

    def set_everything_not_done(self):
        self.is_rating_done = False
        self.is_total_done = False
        self.is_aic_done = False
        self.is_stream_done = False
        self.is_svo_done = False
        self.is_dormitory_done = False
        self.is_orphan_done = False

    def no_data_warning(self, excel):
        info = 'Недостаточно данных для проверки уникальности данных'
        self.parent.logger.warning(f'{info} в таблице {excel}')
        QMessageBox.warning(self.parent, 'Предупреждение',
                            f'{info}\n(возможны дубликаты)')

    def success(self):
        QMessageBox.information(self.parent, "Успешно", "Файлы Excel успешно созданы!")
        done = [f'{'Рейтинг' if self.is_rating_done else ''}',
                f'{'Общий' if self.is_total_done else ''}',
                f'{'АИС' if self.is_aic_done else ''}']
        done.append('Поток') if self.is_stream_done else None
        done.append('СВО') if self.is_svo_done else None
        done.append('Общага') if self.is_dormitory_done else None
        done.append('Сироты') if self.is_orphan_done else None

        self.parent.logger.info('Выполнено корректно: ' + ', '.join(done))
